"""
Excel读取引擎模块
读取标识清单Excel，提取结构化数据
自动识别表头行，返回列名+数据行列表
"""
import os
from typing import List, Dict, Any, Optional
import openpyxl


class ExcelEngine:
    """Excel文件读取引擎类"""

    def read_signage_list(self, file_path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """
        读取标识清单Excel，提取结构化数据。
        自动识别表头行，返回列名+数据行列表。

        Args:
            file_path: Excel文件路径
            sheet_name: 指定sheet名称，为None时取第一个sheet

        Returns:
            包含headers和rows的字典
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            # 选择sheet
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    raise ValueError(f"工作表 '{sheet_name}' 不存在，可用工作表: {wb.sheetnames}")
                ws = wb[sheet_name]
            else:
                ws = wb.active if wb.active else wb[wb.sheetnames[0]]

            # 读取所有行数据
            all_rows: List[List[Any]] = []
            for row in ws.iter_rows(values_only=True):
                all_rows.append(list(row))

            if not all_rows:
                return {"headers": [], "rows": [], "sheet_name": ws.title, "total_rows": 0}

            # 自动识别表头行
            header_row_idx = self._detect_header_row(all_rows)

            headers = [str(cell).strip() if cell is not None else f"列{i+1}" for i, cell in enumerate(all_rows[header_row_idx])]
            rows = []
            for row in all_rows[header_row_idx + 1:]:
                # 跳过完全为空的行
                if all(cell is None or str(cell).strip() == "" for cell in row):
                    continue
                rows.append(row)

            return {
                "headers": headers,
                "rows": rows,
                "sheet_name": ws.title,
                "total_rows": len(rows),
                "header_row": header_row_idx + 1
            }
        finally:
            wb.close()

    def detect_structure(self, file_path: str) -> Dict[str, Any]:
        """
        返回所有sheet名+每个sheet的列名+前3行预览

        Args:
            file_path: Excel文件路径

        Returns:
            包含sheets结构信息的字典
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            sheets_info = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                all_rows: List[List[Any]] = []
                for row in ws.iter_rows(values_only=True):
                    all_rows.append(list(row))

                if not all_rows:
                    sheets_info.append({
                        "name": sheet_name,
                        "headers": [],
                        "preview": [],
                        "total_rows": 0
                    })
                    continue

                header_row_idx = self._detect_header_row(all_rows)
                headers = [str(cell).strip() if cell is not None else f"列{i+1}" for i, cell in enumerate(all_rows[header_row_idx])]

                # 前3行数据预览（不含表头）
                preview_rows = all_rows[header_row_idx + 1:header_row_idx + 4]
                preview = []
                for row in preview_rows:
                    preview.append([str(cell) if cell is not None else "" for cell in row])

                data_rows = [r for r in all_rows[header_row_idx + 1:] if not all(cell is None or str(cell).strip() == "" for cell in r)]

                sheets_info.append({
                    "name": sheet_name,
                    "headers": headers,
                    "preview": preview,
                    "total_rows": len(data_rows)
                })

            return {
                "file_name": os.path.basename(file_path),
                "file_path": file_path,
                "sheets": sheets_info,
                "sheet_count": len(sheets_info)
            }
        finally:
            wb.close()

    def _detect_header_row(self, all_rows: List[List[Any]]) -> int:
        """
        自动检测表头行索引。
        策略：找到第一个有>=2个非空单元格的行。

        Args:
            all_rows: 所有行数据

        Returns:
            表头行的索引（0-based）
        """
        if not all_rows:
            return 0

        for idx, row in enumerate(all_rows[:5]):
            non_empty_count = sum(1 for cell in row if cell is not None and str(cell).strip() != "")
            if non_empty_count >= 2:
                return idx

        return 0
