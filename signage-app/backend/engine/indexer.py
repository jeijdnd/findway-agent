"""
项目索引器模块
扫描本地旧项目目录，建立可检索的项目指纹索引
"""
import os
import json
from datetime import datetime
from typing import List, Dict, Any, Optional

from backend.i18n import _

class ProjectIndexer:
    """项目索引器类，负责扫描和索引旧项目"""
    
    def __init__(self):
        """初始化索引器"""
        self.index_path = os.path.join(os.path.dirname(__file__), "..", "data", "old_projects_index.json")
    
    def scan_directory(self, dir_path: str) -> List[Dict[str, Any]]:
        """
        扫描指定目录，查找所有Excel文件并提取指纹信息
        
        Args:
            dir_path: 要扫描的目录路径
            
        Returns:
            包含项目指纹信息的列表
        """
        projects = []
        
        # 检查目录是否存在
        if not os.path.exists(dir_path):
            print(_("index_dir_not_found", path=dir_path))
            return projects
        
        # 递归查找所有.xlsx文件
        for root, dirs, files in os.walk(dir_path):
            for file in files:
                if file.lower().endswith('.xlsx'):
                    file_path = os.path.join(root, file)
                    project_info = self._extract_project_info(file_path, dir_path)
                    if project_info:
                        projects.append(project_info)
        
        return projects
    
    def _extract_project_info(self, file_path: str, base_dir: str) -> Optional[Dict[str, Any]]:
        """
        从Excel文件中提取项目信息
        
        Args:
            file_path: Excel文件路径
            base_dir: 基础目录路径
            
        Returns:
            项目信息字典，如果提取失败返回None
        """
        try:
            import openpyxl
            
            # 获取相对路径
            relative_path = os.path.relpath(file_path, base_dir)
            
            # 读取Excel文件
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
            # 获取第一个sheet
            sheet_names = wb.sheetnames
            if not sheet_names:
                wb.close()
                return None
            
            first_sheet = wb[sheet_names[0]]
            
            # 读取表头行（假设第一行是表头）
            headers = []
            for cell in first_sheet[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())
            
            # 统计行数（不包括表头）
            row_count = max(0, first_sheet.max_row - 1) if first_sheet.max_row else 0
            
            # 提取文件名（不含扩展名）
            file_name = os.path.splitext(os.path.basename(file_path))[0]
            
            # 分析项目类型（基于文件名和表头）
            project_type = self._analyze_project_type(file_name, headers)
            
            wb.close()
            
            return {
                "id": f"old_{len(file_path)}_{hash(file_path) % 10000}",
                "file_name": file_name,
                "file_path": file_path,
                "relative_path": relative_path,
                "sheet_names": sheet_names,
                "headers": headers,
                "row_count": row_count,
                "project_type": project_type,
                "scan_time": datetime.now().isoformat()
            }
            
        except Exception as e:
            print(_("index_extract_failed", path=file_path, error=e))
            return None
    
    def _analyze_project_type(self, file_name: str, headers: List[str]) -> str:
        """
        分析项目类型
        
        Args:
            file_name: 文件名
            headers: 表头列表
            
        Returns:
            项目类型字符串
        """
        file_name_lower = file_name.lower()
        headers_str = " ".join(headers).lower()
        
        # 基于文件名和表头判断项目类型
        if any(kw in file_name_lower or kw in headers_str for kw in ["学校", "教育", "教学", "校园"]):
            return "学校"
        elif any(kw in file_name_lower or kw in headers_str for kw in ["办公", "写字楼", "商务"]):
            return "办公"
        elif any(kw in file_name_lower or kw in headers_str for kw in ["住宅", "公寓", "居住"]):
            return "住宅"
        elif any(kw in file_name_lower or kw in headers_str for kw in ["医院", "医疗", "卫生"]):
            return "医院"
        elif any(kw in file_name_lower or kw in headers_str for kw in ["商业", "商场", "购物"]):
            return "商业"
        elif any(kw in file_name_lower or kw in headers_str for kw in ["工业", "工厂", "生产"]):
            return "工业"
        else:
            return "其他"
    
    def build_index(self, dir_path: str) -> Dict[str, Any]:
        """
        构建项目索引
        
        Args:
            dir_path: 要扫描的目录路径
            
        Returns:
            索引数据字典
        """
        # 扫描目录获取项目列表
        projects = self.scan_directory(dir_path)
        
        # 构建索引数据
        index_data = {
            "projects": projects,
            "last_scan": datetime.now().isoformat(),
            "scan_directory": dir_path,
            "total_count": len(projects)
        }
        
        # 保存索引文件
        try:
            os.makedirs(os.path.dirname(self.index_path), exist_ok=True)
            with open(self.index_path, "w", encoding="utf-8") as f:
                json.dump(index_data, f, ensure_ascii=False, indent=2)
            print(_("index_saved", path=self.index_path))
        except Exception as e:
            print(_("index_save_failed", error=e))
        
        return index_data
    
    def load_index(self) -> Dict[str, Any]:
        """
        加载已保存的索引
        
        Returns:
            索引数据字典
        """
        try:
            if os.path.exists(self.index_path):
                with open(self.index_path, "r", encoding="utf-8") as f:
                    return json.load(f)
        except Exception as e:
            print(_("index_load_failed", error=e))
        
        return {
            "projects": [],
            "last_scan": None,
            "scan_directory": None,
            "total_count": 0
        }
    
    def search_projects(self, project_type: str = None, keywords: str = None) -> List[Dict[str, Any]]:
        """
        搜索项目
        
        Args:
            project_type: 项目类型
            keywords: 关键词
            
        Returns:
            匹配的项目列表
        """
        index_data = self.load_index()
        projects = index_data.get("projects", [])
        
        results = []
        for project in projects:
            score = self._calculate_match_score(project, project_type, keywords)
            if score > 0:
                project_with_score = project.copy()
                project_with_score["match_score"] = score
                project_with_score["match_reason"] = self._get_match_reason(project, project_type, keywords)
                results.append(project_with_score)
        
        # 按匹配分数降序排序
        results.sort(key=lambda x: x["match_score"], reverse=True)
        
        return results[:5]  # 返回TOP 5
    
    def _calculate_match_score(self, project: Dict[str, Any], project_type: str = None, keywords: str = None) -> float:
        """
        计算项目匹配分数
        
        Args:
            project: 项目信息
            project_type: 项目类型
            keywords: 关键词
            
        Returns:
            匹配分数 (0-100)
        """
        score = 0.0
        
        # 类型匹配 (50%)
        if project_type:
            project_type_str = project.get("project_type", "").lower()
            if project_type.lower() in project_type_str or project_type_str in project_type.lower():
                score += 50.0
        
        # 关键词匹配 (50%)
        if keywords:
            keywords_lower = keywords.lower()
            # 检查文件名
            file_name = project.get("file_name", "").lower()
            if keywords_lower in file_name:
                score += 25.0
            
            # 检查表头
            headers = " ".join(project.get("headers", [])).lower()
            if keywords_lower in headers:
                score += 25.0
        
        # 如果没有指定类型和关键词，给一个基础分
        if not project_type and not keywords:
            score = 10.0
        
        return score
    
    def _get_match_reason(self, project: Dict[str, Any], project_type: str = None, keywords: str = None) -> str:
        """
        生成匹配原因说明
        
        Args:
            project: 项目信息
            project_type: 项目类型
            keywords: 关键词
            
        Returns:
            匹配原因字符串
        """
        reasons = []
        
        if project_type:
            project_type_str = project.get("project_type", "")
            if project_type.lower() in project_type_str.lower() or project_type_str.lower() in project_type.lower():
                reasons.append(f"项目类型匹配：{project_type_str}")
        
        if keywords:
            keywords_lower = keywords.lower()
            file_name = project.get("file_name", "").lower()
            if keywords_lower in file_name:
                reasons.append(f"文件名包含关键词：{keywords}")
            
            headers = " ".join(project.get("headers", [])).lower()
            if keywords_lower in headers:
                reasons.append(f"表头包含关键词：{keywords}")
        
        if not reasons:
            reasons.append("基础匹配")
        
        return "；".join(reasons)