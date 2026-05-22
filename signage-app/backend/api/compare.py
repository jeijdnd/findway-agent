"""
清单对比API模块
提供清单文件上传、对比和导出功能
"""
from fastapi import APIRouter, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Dict, Any, Optional
import os
import io
import json
import tempfile

from backend.engine.excel_engine import ExcelEngine
from backend.engine.differ import ListDiffer

router = APIRouter()

excel_engine = ExcelEngine()
differ = ListDiffer()


class CompareRequest(BaseModel):
    """对比请求模型"""
    file_path_a: str
    file_path_b: str
    key_field: str = "编号"
    sheet_name_a: Optional[str] = None
    sheet_name_b: Optional[str] = None


class StructureResponse(BaseModel):
    """文件结构响应模型"""
    file_name: str
    file_path: str
    sheets: List[Dict[str, Any]]
    sheet_count: int


class CompareResultResponse(BaseModel):
    """对比结果响应模型"""
    added: List[Dict[str, Any]]
    removed: List[Dict[str, Any]]
    modified: List[Dict[str, Any]]
    unchanged: List[Dict[str, Any]]
    summary: Dict[str, Any]


@router.post("/api/compare/upload")
async def upload_file(file: UploadFile = File(...)):
    """
    上传Excel文件到临时目录，返回文件路径供后续对比使用
    """
    if not file.filename or not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="只支持 .xlsx / .xls 文件")

    try:
        # 保存到临时目录
        temp_dir = tempfile.mkdtemp(prefix="findway_compare_")
        file_path = os.path.join(temp_dir, file.filename)
        content = await file.read()
        with open(file_path, "wb") as f:
            f.write(content)

        # 检测文件结构
        structure = excel_engine.detect_structure(file_path)

        return {
            "success": True,
            "file_path": file_path,
            "file_name": structure["file_name"],
            "sheets": structure["sheets"],
            "sheet_count": structure["sheet_count"]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"文件上传失败: {str(e)}")


@router.post("/api/compare/structure")
async def get_structure(file_path: str):
    """
    获取Excel文件的结构信息（sheet名+列名+预览）
    """
    try:
        structure = excel_engine.detect_structure(file_path)
        return structure
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"读取文件结构失败: {str(e)}")


@router.post("/api/compare/run", response_model=CompareResultResponse)
async def run_compare(request: CompareRequest):
    """
    执行清单对比，返回差异结果
    """
    try:
        # 读取两个文件
        data_a = excel_engine.read_signage_list(request.file_path_a, request.sheet_name_a)
        data_b = excel_engine.read_signage_list(request.file_path_b, request.sheet_name_b)

        # 执行对比
        result = differ.compare(
            list_a=data_a["rows"],
            list_b=data_b["rows"],
            headers_a=data_a["headers"],
            headers_b=data_b["headers"],
            key_field=request.key_field
        )

        # 附加文件信息
        result["summary"]["file_a"] = os.path.basename(request.file_path_a)
        result["summary"]["file_b"] = os.path.basename(request.file_path_b)
        result["summary"]["headers_a"] = data_a["headers"]
        result["summary"]["headers_b"] = data_b["headers"]

        return result
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"对比失败: {str(e)}")


@router.post("/api/compare/export")
async def export_compare(compare_data: str = Form(...)):
    """
    导出差异对比报告为Excel文件
    compare_data: JSON格式的对比结果
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        data = json.loads(compare_data)
        summary = data.get("summary", {})

        wb = openpyxl.Workbook()

        # 样式定义
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=12, color="FFFFFF")
        added_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 绿色
        removed_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 红色
        modified_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # 黄色
        center_align = Alignment(horizontal="center", vertical="center")

        # === 汇总页 ===
        ws_summary = wb.active
        ws_summary.title = "对比汇总"
        ws_summary.append(["清单差异对比报告"])
        ws_summary.append([])
        ws_summary.append(["旧版文件", summary.get("file_a", "")])
        ws_summary.append(["新版文件", summary.get("file_b", "")])
        ws_summary.append(["匹配字段", summary.get("key_field", "")])
        ws_summary.append([])
        ws_summary.append(["统计项", "数量"])
        ws_summary.append(["旧版行数", summary.get("total_old", 0)])
        ws_summary.append(["新版行数", summary.get("total_new", 0)])
        ws_summary.append(["新增项", summary.get("added_count", 0)])
        ws_summary.append(["删除项", summary.get("removed_count", 0)])
        ws_summary.append(["修改项", summary.get("modified_count", 0)])
        ws_summary.append(["不变项", summary.get("unchanged_count", 0)])

        # === 新增项 ===
        added = data.get("added", [])
        if added:
            ws_added = wb.create_sheet("新增项")
            headers = summary.get("headers_b", [])
            ws_added.append(["编号"] + headers)
            for row_idx, item in enumerate(added, start=2):
                row_data = [item.get("key", "")]
                for h in headers:
                    row_data.append(item.get("data", {}).get(h, ""))
                ws_added.append(row_data)
                # 设置绿色背景
                for col in range(1, len(row_data) + 1):
                    ws_added.cell(row=row_idx, column=col).fill = added_fill

        # === 删除项 ===
        removed = data.get("removed", [])
        if removed:
            ws_removed = wb.create_sheet("删除项")
            headers = summary.get("headers_a", [])
            ws_removed.append(["编号"] + headers)
            for row_idx, item in enumerate(removed, start=2):
                row_data = [item.get("key", "")]
                for h in headers:
                    row_data.append(item.get("data", {}).get(h, ""))
                ws_removed.append(row_data)
                for col in range(1, len(row_data) + 1):
                    ws_removed.cell(row=row_idx, column=col).fill = removed_fill

        # === 修改项 ===
        modified = data.get("modified", [])
        if modified:
            ws_modified = wb.create_sheet("修改项")
            headers = list(dict.fromkeys(
                summary.get("headers_a", []) + summary.get("headers_b", [])
            ))
            ws_modified.append(["编号", "修改字段", "旧值", "新值"])
            for row_idx, item in enumerate(modified, start=2):
                for change in item.get("changed_fields", []):
                    ws_modified.append([
                        item.get("key", ""),
                        change.get("field", ""),
                        change.get("old_value", ""),
                        change.get("new_value", "")
                    ])
                    ws_modified.cell(row=row_idx, column=1).fill = modified_fill
                    ws_modified.cell(row=row_idx, column=2).fill = modified_fill
                    row_idx += 1

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        wb.close()

        filename = f"清单对比报告_{summary.get('file_a', '')}_vs_{summary.get('file_b', '')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")
