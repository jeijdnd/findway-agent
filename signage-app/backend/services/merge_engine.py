"""
兔钉清单合并引擎
读取兔钉导出 xlsx + 完整清单模板，按编号精确匹配并生成新清单
"""

import os
import shutil
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.workbook.workbook import Workbook

# 完整清单中应跳过的 Sheet 名称关键词
SKIP_SHEET_KEYWORDS = ["00材料", "汇总", "总计", "封面", "目录", "说明"]

# 分类 Sheet 名称关键词（包含任一即为有效分类 Sheet）
CATEGORY_SHEET_KEYWORDS = [
    "户外", "室外", "办公", "商业", "商场", "地下室", "室内",
    "A类", "B类", "C类", "D类",
]

# 新增款式编号前缀 → Sheet 匹配关键词
CODE_PREFIX_SHEET_MAP = {
    "A": ["户外", "室外", "A类"],
    "B": ["室内", "办公", "B类"],
    "C": ["商场", "商业", "C类"],
    "D": ["地下室", "D类"],
}

# 列名候选（模糊匹配）
COL_SEQ = ["序号"]
COL_CODE = ["标识编号", "编号", "图例编号"]
COL_NAME = ["标识名称", "名称", "图例名称"]
COL_QTY = ["数量", "合计"]
COL_MATERIAL = ["材料", "参考材质", "材质"]
COL_SIZE = ["尺寸", "参考尺寸", "规格"]
COL_POWERED = ["是否带电", "带电"]

# 新增行黄色背景
NEW_ITEM_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


class MergeEngine:
    """兔钉清单合并引擎"""

    # ------------------------------------------------------------------ #
    #  读取兔钉导出
    # ------------------------------------------------------------------ #

    def read_tuding_export(self, file_path: str) -> Dict[str, Any]:
        """
        读取兔钉导出 xlsx，解析"图例统计"和"信息" Sheet。

        Returns:
            {project_name, plans, items, source_path}
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")

        wb = openpyxl.load_workbook(file_path, data_only=True)
        try:
            stats_sheet = self._find_sheet(wb, "图例统计")
            if not stats_sheet:
                raise ValueError(f"未找到「图例统计」Sheet，可用: {wb.sheetnames}")

            info_sheet = self._find_sheet(wb, "信息")
            project_name, info_plans = self._parse_info_sheet(info_sheet) if info_sheet else ("", [])

            items, plan_headers = self._parse_stats_sheet(stats_sheet)
            # 优先使用图例统计表头的图纸列名，信息 Sheet 作为补充
            plans = plan_headers if plan_headers else info_plans

            return {
                "project_name": project_name,
                "plans": plans,
                "items": items,
                "source_path": file_path,
            }
        finally:
            wb.close()

    def _parse_info_sheet(self, ws) -> Tuple[str, List[str]]:
        """解析「信息」Sheet，提取项目名称和图纸列表"""
        project_name = ""
        plans: List[str] = []

        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if not any(cells):
                continue
            # 查找项目名称
            for i, cell in enumerate(cells):
                if "项目名称" in cell or "工程名称" in cell:
                    if i + 1 < len(cells) and cells[i + 1]:
                        project_name = cells[i + 1]
                    break
                if "项目" in cell and i + 1 < len(cells) and cells[i + 1] and not project_name:
                    project_name = cells[i + 1]
            # 查找图纸/楼层列表行
            for i, cell in enumerate(cells):
                if any(kw in cell for kw in ["图纸", "楼层", "平面"]):
                    plans.extend([c for c in cells[i + 1:] if c and c not in plans])

        return project_name, plans

    def _parse_stats_sheet(self, ws) -> Tuple[List[Dict], List[str]]:
        """
        解析「图例统计」Sheet。
        前三列为固定字段（序号/编号/名称），其余列为动态数量列。
        """
        all_rows: List[List] = []
        for row in ws.iter_rows(values_only=True):
            all_rows.append(list(row))

        header_idx = self._detect_stats_header(all_rows)
        header_row = all_rows[header_idx]
        headers = [str(c).strip() if c is not None else "" for c in header_row]

        # 动态数量列：跳过前三列（序号/编号/名称）
        plan_headers = [h for h in headers[3:] if h]
        qty_col_start = 3

        items: List[Dict] = []
        for row in all_rows[header_idx + 1:]:
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue

            seq = row[0] if len(row) > 0 else ""
            code = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            name = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""

            if not code:
                continue

            plan_quantities: Dict[str, float] = {}
            total_qty = 0.0
            for col_idx in range(qty_col_start, len(row)):
                plan_name = headers[col_idx] if col_idx < len(headers) else f"列{col_idx + 1}"
                if not plan_name:
                    continue
                qty = self._to_number(row[col_idx])
                if qty > 0:
                    plan_quantities[plan_name] = qty
                    total_qty += qty

            items.append({
                "seq": seq,
                "code": code,
                "name": name,
                "total_qty": total_qty,
                "plan_quantities": plan_quantities,
            })

        return items, plan_headers

    def _detect_stats_header(self, all_rows: List[List]) -> int:
        """检测图例统计表头行（含序号/编号/名称关键词）"""
        for idx, row in enumerate(all_rows[:10]):
            cells = [str(c).strip() if c is not None else "" for c in row]
            has_seq = any(k in c for c in cells for k in COL_SEQ)
            has_code = any(k in c for c in cells for k in COL_CODE)
            has_name = any(k in c for c in cells for k in COL_NAME)
            if has_seq and (has_code or has_name):
                return idx
        return 0

    # ------------------------------------------------------------------ #
    #  读取完整清单
    # ------------------------------------------------------------------ #

    def read_full_checklist(self, file_path: str) -> Dict[str, Any]:
        """
        读取完整清单（和阳楼格式），建立编号索引。

        Returns:
            {source_path, index, category_sheets, sheet_meta, workbook_path}
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")

        wb = openpyxl.load_workbook(file_path, data_only=True)
        index: Dict[str, Dict] = {}
        category_sheets: List[str] = []
        sheet_meta: Dict[str, Dict] = {}

        try:
            for sheet_name in wb.sheetnames:
                if self._should_skip_sheet(sheet_name):
                    continue
                if not self._is_category_sheet(sheet_name):
                    continue

                ws = wb[sheet_name]
                sheet_index, meta = self._parse_checklist_sheet(ws, sheet_name)
                category_sheets.append(sheet_name)
                sheet_meta[sheet_name] = meta
                for code, item in sheet_index.items():
                    if code in index:
                        # 编号重复时保留首次出现
                        continue
                    index[code] = item
        finally:
            wb.close()

        return {
            "source_path": file_path,
            "index": index,
            "category_sheets": category_sheets,
            "sheet_meta": sheet_meta,
            "workbook_path": file_path,
        }

    def _should_skip_sheet(self, name: str) -> bool:
        """判断是否跳过该 Sheet"""
        return any(kw in name for kw in SKIP_SHEET_KEYWORDS)

    def _is_category_sheet(self, name: str) -> bool:
        """判断是否为分类 Sheet"""
        return any(kw in name for kw in CATEGORY_SHEET_KEYWORDS)

    def _parse_checklist_sheet(self, ws, sheet_name: str) -> Tuple[Dict[str, Dict], Dict]:
        """解析单个分类 Sheet，返回编号索引和 Sheet 元数据"""
        all_rows: List[List] = []
        for row in ws.iter_rows(values_only=True):
            all_rows.append(list(row))

        if not all_rows:
            return {}, {"header_row": 0, "cols": {}}

        header_idx = self._detect_checklist_header(all_rows)
        header_row = all_rows[header_idx]
        headers = [str(c).strip() if c is not None else "" for c in header_row]

        cols = {
            "seq": self._find_col(headers, COL_SEQ),
            "code": self._find_col(headers, COL_CODE),
            "name": self._find_col(headers, COL_NAME),
            "qty": self._find_col(headers, COL_QTY),
            "material": self._find_col(headers, COL_MATERIAL),
            "size": self._find_col(headers, COL_SIZE),
            "powered": self._find_col(headers, COL_POWERED),
        }

        sheet_index: Dict[str, Dict] = {}
        for row_num, row in enumerate(all_rows[header_idx + 1:], start=header_idx + 2):
            code = self._cell_str(row, cols["code"])
            if not code:
                continue

            sheet_index[code] = {
                "sheet": sheet_name,
                "row": row_num,
                "seq": self._cell_str(row, cols["seq"]),
                "code": code,
                "name": self._cell_str(row, cols["name"]),
                "qty": self._to_number(self._cell_val(row, cols["qty"])),
                "material": self._cell_str(row, cols["material"]),
                "size": self._cell_str(row, cols["size"]),
                "powered": self._cell_str(row, cols["powered"]),
            }

        meta = {"header_row": header_idx + 1, "cols": cols, "headers": headers}
        return sheet_index, meta

    def _detect_checklist_header(self, all_rows: List[List]) -> int:
        """检测完整清单表头行"""
        for idx, row in enumerate(all_rows[:10]):
            headers = [str(c).strip() if c is not None else "" for c in row]
            if self._find_col(headers, COL_CODE) >= 0:
                return idx
        return 0

    # ------------------------------------------------------------------ #
    #  合并逻辑
    # ------------------------------------------------------------------ #

    def merge(self, tuding_data: Dict, checklist_data: Dict) -> Dict[str, Any]:
        """
        按编号精确匹配兔钉导出 vs 完整清单。

        Returns:
            MergeResult 字典
        """
        tuding_items = {item["code"]: item for item in tuding_data.get("items", [])}
        checklist_index = checklist_data.get("index", {})
        source_template = checklist_data.get("source_path", "")

        matches: List[Dict] = []
        new_items: List[Dict] = []
        missing_items: List[Dict] = []

        matched_codes = set()

        # 遍历兔钉项：匹配 or 新增
        for code, t_item in tuding_items.items():
            if code in checklist_index:
                c_item = checklist_index[code]
                matches.append({
                    "code": code,
                    "name": t_item.get("name") or c_item.get("name", ""),
                    "status": "matched",
                    "old_qty": c_item.get("qty", 0),
                    "new_qty": t_item.get("total_qty", 0),
                    "sheet_name": c_item.get("sheet", ""),
                    "row": c_item.get("row", 0),
                    "plan_quantities": t_item.get("plan_quantities", {}),
                })
                matched_codes.add(code)
            else:
                new_items.append({
                    "code": code,
                    "name": t_item.get("name", ""),
                    "status": "new_item",
                    "old_qty": 0,
                    "new_qty": t_item.get("total_qty", 0),
                    "plan_quantities": t_item.get("plan_quantities", {}),
                })

        # 遍历清单项：兔钉没有的 → 归零
        for code, c_item in checklist_index.items():
            if code not in tuding_items:
                entry = {
                    "code": code,
                    "name": c_item.get("name", ""),
                    "status": "zeroed",
                    "old_qty": c_item.get("qty", 0),
                    "new_qty": 0,
                    "sheet_name": c_item.get("sheet", ""),
                    "row": c_item.get("row", 0),
                }
                matches.append(entry)
                missing_items.append(entry)

        total_matched = sum(1 for m in matches if m["status"] == "matched")
        total_new = len(new_items)
        total_zeroed = len(missing_items)

        return {
            "project_name": tuding_data.get("project_name", ""),
            "plans": tuding_data.get("plans", []),
            "matches": matches,
            "new_items": new_items,
            "missing_items": missing_items,
            "source_template": source_template,
            "total_matched": total_matched,
            "total_new": total_new,
            "total_zeroed": total_zeroed,
        }

    # ------------------------------------------------------------------ #
    #  生成清单
    # ------------------------------------------------------------------ #

    def generate_checklist(
        self,
        merge_result: Dict,
        template_path: str,
        output_path: str,
        checklist_data: Optional[Dict] = None,
    ) -> str:
        """
        复制模板 → 更新数量 → 追加新增项 → 保存到 output_path。
        永远不修改原始模板文件。
        """
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"模板文件不存在: {template_path}")

        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        shutil.copy2(template_path, output_path)

        wb = openpyxl.load_workbook(output_path)
        sheet_meta = (checklist_data or {}).get("sheet_meta", {})
        category_sheets = (checklist_data or {}).get("category_sheets", [])

        # 更新已匹配项和归零项的数量
        for match in merge_result.get("matches", []):
            sheet_name = match.get("sheet_name")
            row = match.get("row")
            if not sheet_name or not row or sheet_name not in wb.sheetnames:
                continue
            meta = sheet_meta.get(sheet_name, {})
            qty_col = meta.get("cols", {}).get("qty", -1)
            if qty_col < 0:
                continue
            ws = wb[sheet_name]
            ws.cell(row=row, column=qty_col + 1, value=match.get("new_qty", 0))

        # 追加新增项
        for new_item in merge_result.get("new_items", []):
            target_sheet = self._resolve_target_sheet(
                new_item.get("code", ""),
                category_sheets,
                wb,
            )
            if not target_sheet:
                continue
            self._append_new_row(wb[target_sheet], new_item, sheet_meta.get(target_sheet, {}))

        # 更新汇总/总计 Sheet 公式（如果存在）
        self._refresh_summary_sheets(wb)

        wb.save(output_path)
        wb.close()
        return output_path

    def _resolve_target_sheet(
        self, code: str, category_sheets: List[str], wb: Workbook
    ) -> Optional[str]:
        """根据编号前缀确定目标 Sheet"""
        prefix = code[0].upper() if code else ""
        keywords = CODE_PREFIX_SHEET_MAP.get(prefix, [])

        if keywords:
            for sheet_name in category_sheets:
                if sheet_name in wb.sheetnames and any(kw in sheet_name for kw in keywords):
                    return sheet_name

        # 其他编号 → 追加到最后一个分类 Sheet
        for sheet_name in reversed(category_sheets):
            if sheet_name in wb.sheetnames:
                return sheet_name
        return None

    def _append_new_row(self, ws, new_item: Dict, meta: Dict):
        """在 Sheet 末尾追加新增款式行，黄色背景标注"""
        cols = meta.get("cols", {})
        last_row = ws.max_row
        new_row = last_row + 1

        col_map = {
            cols.get("seq", 0): new_row - meta.get("header_row", 1),
            cols.get("code", 1): new_item.get("code", ""),
            cols.get("name", 2): new_item.get("name", ""),
            cols.get("qty", 3): new_item.get("new_qty", 0),
        }
        for col_idx, value in col_map.items():
            if col_idx >= 0:
                cell = ws.cell(row=new_row, column=col_idx + 1, value=value)
                cell.fill = NEW_ITEM_FILL

        max_col = max(cols.values()) + 1 if cols else ws.max_column
        for col in range(1, max_col + 1):
            ws.cell(row=new_row, column=col).fill = NEW_ITEM_FILL

    def _refresh_summary_sheets(self, wb: Workbook):
        """尝试刷新汇总/总计 Sheet（触发公式重算标记）"""
        for sheet_name in wb.sheetnames:
            if any(kw in sheet_name for kw in ["汇总", "总计"]):
                ws = wb[sheet_name]
                # openpyxl 不自动重算公式，标记 sheet 为已修改以提示 Excel/WPS 打开时重算
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                            cell.value = cell.value  # 保持公式不变，触发 dirty 标记

    # ------------------------------------------------------------------ #
    #  模板扫描
    # ------------------------------------------------------------------ #

    def list_templates(self, scan_dirs: Optional[List[str]] = None) -> List[Dict]:
        """
        扫描目录下所有完整清单 xlsx 模板。
        返回 [{name, path, sheets}]
        """
        if scan_dirs is None:
            base = os.path.dirname(os.path.dirname(__file__))
            scan_dirs = [
                os.path.join(base, "data", "templates"),
                os.path.join(base, "data"),
            ]

        templates: List[Dict] = []
        seen_paths: set = set()

        for scan_dir in scan_dirs:
            if not os.path.isdir(scan_dir):
                continue
            for root, _, files in os.walk(scan_dir):
                for fname in files:
                    if not fname.lower().endswith((".xlsx", ".xls")):
                        continue
                    fpath = os.path.normpath(os.path.join(root, fname))
                    if fpath in seen_paths:
                        continue
                    seen_paths.add(fpath)
                    try:
                        sheets = self._get_category_sheet_names(fpath)
                        if sheets:
                            templates.append({
                                "name": fname,
                                "path": fpath,
                                "sheets": sheets,
                            })
                    except Exception:
                        continue

        return templates

    def _get_category_sheet_names(self, file_path: str) -> List[str]:
        """获取文件中所有分类 Sheet 名称"""
        wb = openpyxl.load_workbook(file_path, read_only=True)
        try:
            return [
                s for s in wb.sheetnames
                if not self._should_skip_sheet(s) and self._is_category_sheet(s)
            ]
        finally:
            wb.close()

    # ------------------------------------------------------------------ #
    #  工具方法
    # ------------------------------------------------------------------ #

    def _find_sheet(self, wb: Workbook, keyword: str):
        """按关键词查找 Sheet（精确或包含）"""
        if keyword in wb.sheetnames:
            return wb[keyword]
        for name in wb.sheetnames:
            if keyword in name:
                return wb[name]
        return None

    def _find_col(self, headers: List[str], candidates: List[str]) -> int:
        """在表头中查找列索引（-1 表示未找到）"""
        for i, h in enumerate(headers):
            for c in candidates:
                if c in h:
                    return i
        return -1

    def _cell_val(self, row: List, col_idx: int):
        """安全取单元格值"""
        if col_idx < 0 or col_idx >= len(row):
            return None
        return row[col_idx]

    def _cell_str(self, row: List, col_idx: int) -> str:
        """安全取单元格字符串"""
        val = self._cell_val(row, col_idx)
        return str(val).strip() if val is not None else ""

    def _to_number(self, val) -> float:
        """将单元格值转为数字"""
        if val is None:
            return 0.0
        if isinstance(val, (int, float)):
            return float(val)
        try:
            return float(str(val).strip())
        except (ValueError, TypeError):
            return 0.0


# 全局实例
merge_engine = MergeEngine()
